"""Модуль сохранения результатов тестовых сессий в XLSX."""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl

import config

if TYPE_CHECKING:
    from test_session import Test


def save_results(test: Test) -> None:
    """Сохраняет результат сессии в таблицу."""
    try:
        if not Path(config.RESULTS_FILE).exists():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append([
                "№",
                "участник",
                "школа",
                "верно",
                "ошибки",
                f"штрафы (ошибки × {config.PUNISH_FACTOR})",
                "время (сек)",
                "статус",
            ])
            wb.save(config.RESULTS_FILE)

        wb = openpyxl.load_workbook(config.RESULTS_FILE)
        ws = wb.active
        next_number = ws.max_row

        ws.append([
            next_number,
            test.student_name,
            test.student_school,
            test.right_answers,
            test.wrong_answers,
            test.punish_score,
            test.time_taken,
            test.status,
        ])

        wb.save(config.RESULTS_FILE)
    except Exception as e:
        print(f"Error saving results: {e}")
