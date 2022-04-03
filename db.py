import pandas as pd
import os
from openpyxl import load_workbook
import datetime


class TEST():
    def __init__(self, file):
        self.questions = []
        self.results = []
        self.df = pd.read_excel(file)
        self.df = self.df.dropna(how='all').dropna(how='all', axis=1)
        self.make_questions()

    def make_questions(self):
        for index, row in self.df.iterrows():
            # игнорируем пустые ответы №4, иначе pd запишет туда NaN
            # нужно ли игнорировать пустой текст вопроса?
            if isinstance(row[6], str):
                ans = row[3:7]
            else:
                ans = row[3:6]

            question = Question(
                number=row["№"],
                right_answer=row[1],
                image=f'assets/{row["№"]}.png',
                text=row[2],
                answers=ans
            )
            self.questions.append(question)

    def write_user_results(
        self,
        results_file="результаты.xlsx",
        first_name="имя",
        last_name="фамилия",
        time_total=datetime.timedelta(seconds=10),
        wrong_answers="неизвестно"
    ):

        user_data = [
            last_name,
            first_name,
            round(time_total.total_seconds(), 1),
            wrong_answers
        ]

        columns_names = [
            "фамилия",
            "имя",
            "время (с)",
            "ошибки"
        ]

        if not os.path.isfile(results_file):
            user_df = pd.DataFrame(
                [user_data],
                columns=columns_names,
                index=[1]
            )
            user_df.to_excel(results_file)
        else:
            with pd.ExcelWriter(
                results_file,
                mode="a",
                engine="openpyxl",
                if_sheet_exists="overlay"
            ) as writer:
                user_df = pd.DataFrame(
                    [user_data],
                    columns=columns_names,
                    index=[writer.sheets['Sheet1'].max_row]
                )
                user_df.to_excel(
                    writer,
                    index_label="№",
                    sheet_name="Sheet1",
                    startrow=writer.sheets['Sheet1'].max_row,
                    header=False
                )


class Question:
    def __init__(
        self,
        number=0,
        right_answer=0,
        text="",
        image="assets/01.png",
        answers=["варианты ответа не созданы"]
    ):
        self.number = number
        self.right_answer = right_answer
        self.text = text
        self.image = image
        self.answers = answers

if __name__ == "__main__":
    test = TEST("assets/db.xlsx")
    test.write_user_results(
        first_name="Второй",
        last_name="Питонов",
        wrong_answers=10
    )
